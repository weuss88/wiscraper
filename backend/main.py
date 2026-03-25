import asyncio
import csv
import io
import json
import uuid
from datetime import datetime

import pandas as pd
from fastapi import BackgroundTasks, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse

from database import (create_job, delete_job, get_job, get_results, init_db,
                      save_result, update_job)
from models import ScrapeRequest
from scrapers.maps_scraper import run_maps_scraper
from scrapers.leboncoin_scraper import run_leboncoin_scraper
from scrapers.pap_scraper import run_pap_scraper

app = FastAPI(title="WiScrap API", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup():
    await init_db()


# ── Helper ────────────────────────────────────────────────
async def run_job(job_id: str, request: ScrapeRequest):
    async def progress_cb(done: int, total: int, msg: str = ""):
        await update_job(job_id, "running", done, total, msg)

    async def save_cb(lead: dict):
        await save_result(job_id, lead)

    try:
        await update_job(job_id, "running", 0, 0, "Démarrage...")

        if request.source in ("maps_fr", "maps_be"):
            await run_maps_scraper(request.maps, progress_cb, save_cb)
        elif request.source == "leboncoin":
            await run_leboncoin_scraper(request.leboncoin, progress_cb, save_cb)
        elif request.source == "pap":
            await run_pap_scraper(request.pap, progress_cb, save_cb)
        else:
            raise ValueError(f"Source inconnue : {request.source}")

        results = await get_results(job_id)
        await update_job(job_id, "done", len(results), len(results), "Terminé ✓")

    except Exception as e:
        await update_job(job_id, "error", 0, 0, str(e))


# ── Routes ────────────────────────────────────────────────
@app.post("/api/scrape")
async def start_scrape(request: ScrapeRequest, background_tasks: BackgroundTasks):
    job_id = str(uuid.uuid4())
    created_at = datetime.now().isoformat()
    await create_job(job_id, request.source, created_at)
    background_tasks.add_task(run_job, job_id, request)
    return {"job_id": job_id}


@app.get("/api/jobs/{job_id}")
async def get_job_status(job_id: str):
    job = await get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job introuvable")
    return job


@app.get("/api/results/{job_id}")
async def get_job_results(job_id: str):
    job = await get_job(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job introuvable")
    results = await get_results(job_id)
    return {"results": results, "total": len(results)}


@app.get("/api/export/{job_id}/{fmt}")
async def export_results(job_id: str, fmt: str):
    results = await get_results(job_id)
    if not results:
        raise HTTPException(status_code=404, detail="Aucun résultat")

    if fmt == "json":
        content = json.dumps(results, ensure_ascii=False, indent=2)
        return Response(
            content=content,
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=wiscrap_{job_id[:8]}.json"},
        )

    if fmt == "csv":
        output = io.StringIO()
        if results:
            writer = csv.DictWriter(output, fieldnames=results[0].keys())
            writer.writeheader()
            writer.writerows(results)
        return Response(
            content=output.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=wiscrap_{job_id[:8]}.csv"},
        )

    if fmt == "xlsx":
        df = pd.DataFrame(results)
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Leads")
            ws = writer.sheets["Leads"]
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
            for cell in ws[1]:
                cell.fill = PatternFill("solid", fgColor="0a3d62")
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.alignment = Alignment(horizontal="center")
            for i, row in enumerate(ws.iter_rows(min_row=2)):
                fill = PatternFill("solid", fgColor="f4f9f4" if i % 2 == 0 else "FFFFFF")
                for cell in row:
                    cell.fill = fill
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
            for i, col in enumerate(df.columns, 1):
                ws.column_dimensions[get_column_letter(i)].width = min(
                    max(len(str(col)) + 4, 12), 40
                )
            ws.freeze_panes = "A2"
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=wiscrap_{job_id[:8]}.xlsx"},
        )

    raise HTTPException(status_code=400, detail="Format invalide (csv|json|xlsx)")


@app.delete("/api/jobs/{job_id}")
async def remove_job(job_id: str):
    await delete_job(job_id)
    return {"ok": True}


@app.get("/health")
async def health():
    return {"status": "ok"}
